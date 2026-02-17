# =========================
# GENERATOR
# =========================

from openpyxl import load_workbook
from datetime import timedelta
import shutil
from infos import INFOS
from vilt import DATE_RANGE, SHEETS, parse_date
from logger import error, log
from config import PATHS
from excel_utils import copy_formulas, copy_style

def fill_resource(ws, name, activities):
    info = INFOS[name]

    ws['B1'] = DATE_RANGE['Start Date']
    ws['C1'] = DATE_RANGE['End Date']
    ws['B4'] = name
    ws['A7'] = info['registration']
    ws['B7'] = info['email']
    ws['C7'] = info['profile']
    ws['D1'] = info['pm']

    row = 10
    last_date = None

    for a in activities:
        ws.cell(row, 1, a['date'])
        #ws.cell(row, 2, a['weekday']) # Iremos seguir a fórmula do Excel nessa coluna
        ws.cell(row, 3, a['hours'])
        ws.cell(row, 4, 'Sim' if a['overwork'] else 'Não')
        #ws.cell(row, 5, a['hours'] * 2 if a['overwork'] else a['hours']) # Iremos seguir a fórmula do Excel nessa coluna
        ws.cell(row, 6, a['activity'])
        last_date = parse_date(a['date'])
        row += 1

    if not last_date:
        last_date = DATE_RANGE['Start Date'] - timedelta(days=1)

    current_date = last_date + timedelta(days=1)
    end_date = DATE_RANGE['End Date']

    while current_date <= end_date:

        if current_date.weekday() < 5:  # 0=segunda, 6=domingo

            ws.cell(row, 1, current_date.strftime("%d/%m/%Y"))
            ws.cell(row, 3, 8)
            ws.cell(row, 4, 'Não')
            ws.cell(row, 5, 8)
            ws.cell(row, 6, 'Previsão de horas')

            row += 1

        current_date += timedelta(days=1)

    return info['profile'], row

def fill_header(ws, info):
    ws['B5'] = DATE_RANGE['Start Date']
    ws['C5'] = DATE_RANGE['End Date']
    ws['B10'] = info['cost_center']
    ws['B12'] = info['coordinator']

def fill_header_pms(ws, pms):
    cell = ws['B7']
    
    if not pms:
        return
    
    if len(pms) == 1:
        cell.value = pms[0]
    else:
        cell.value = " | ".join(pms)


def fill_total_infos(ws, info):
    ws['K3'] = info['cost_center']
    ws['O3'] = info['coordinator']  
    ws['A3'] = info['project']   

def fill_total_hours(ws, name, profile):

    sheet_reference = f"'{name}'!D7"

    for row in range(3, 8):

        profile_cell = ws.cell(row, 4)  # Coluna D
        hours_cell = ws.cell(row, 8)    # Coluna H

        # 🔹 Caso 1: perfil já existe → somar referência
        if profile_cell.value == profile:

            if hours_cell.value:
                hours_cell.value = f"{hours_cell.value}+{sheet_reference}"
            else:
                hours_cell.value = f"={sheet_reference}"

            return

        # 🔹 Caso 2: primeira linha vazia ("-")
        elif profile_cell.value == "-":

            profile_cell.value = profile
            hours_cell.value = f"={sheet_reference}"

            return

def group_by_squad():
    squads = {}
    for name, info in INFOS.items():
        squads.setdefault(info['squad'], []).append(name)
    return squads

def create_files():
    squads = group_by_squad()

    log(f"{len(squads)} squads identificadas")

    for squad, members in squads.items():
        log(f"Gerando planilha da squad: {squad}")

        output = f"{PATHS['output']}/{squad}.xlsx"
        shutil.copyfile(PATHS['template'], output)

        wb = load_workbook(output)

        if SHEETS['base'] not in wb.sheetnames:
            error("Aba base 'sheet' não encontrada no template")

        base_ws = wb[SHEETS['base']]

        pms = {INFOS[name].get('pm') for name in members if INFOS[name].get('pm')}
        fill_header(wb[SHEETS['header']], INFOS[members[0]])
        fill_header_pms(wb[SHEETS['header']], list(pms))

        fill_total_infos(wb[SHEETS['total']], INFOS[members[0]])
        
        for name in members:
            ws = wb.copy_worksheet(base_ws)
            ws.title = name

            activities = [
                d for days in INFOS[name]['data'].values() for d in days
            ]

            profile, total_rows = fill_resource(ws, name, activities)
            copy_style(ws, total_rows)
            copy_formulas(ws, 10, total_rows)
            fill_total_hours(wb[SHEETS['total']], name, profile)


        wb.remove(base_ws)
        wb.save(output)

    log("Processo finalizado com sucesso")