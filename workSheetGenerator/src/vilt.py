# =========================
# VILT ZEUS REPORT
# =========================

from logger import error, log, warn
from datetime import datetime, date
from openpyxl import load_workbook
from infos import INFOS
from config import PATHS

SHEETS = {
    'base': 'sheet',
    'header': 'CABEÇALHO',
    'total': 'TOTALIZADOR'
}

VILT_SHEET_FALLBACK = 'Sheet1'

ROW_IDENTIFIER = ['project', 'company']

VILT_COLUMNS = {
    'Person': None,
    'Day': None,
    'Hours': None,
    'Overwork': None,
    'Notes': None
}

DATE_RANGE = {'Start Date': '', 'End Date': ''}

def find_vilt_sheet(wb):
    if VILT_SHEET_FALLBACK in wb.sheetnames:
        return wb[VILT_SHEET_FALLBACK]
    return wb.active

def find_table_vilt(ws):
    log("Localizando tabela no VILT")

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue

        key = str(row[0]).strip().lower()

        if key.startswith('start'):
            start_date = parse_date(row[1])
            DATE_RANGE['Start Date'] = start_date

            year = start_date.year
            month = start_date.month

            if month == 12:
                next_month = 1
                year += 1
            else:
                next_month = month + 1

            DATE_RANGE['End Date'] = parse_date(date(year, next_month, 20))

        if len(row) > 1:
            if (
                str(row[0]).strip().lower() == ROW_IDENTIFIER[0]
                and str(row[1]).strip().lower() == ROW_IDENTIFIER[1]
            ):
                for idx, col in enumerate(row):
                    if col in VILT_COLUMNS:
                        VILT_COLUMNS[col] = idx

                if None in VILT_COLUMNS.values():
                    error("Colunas obrigatórias não encontradas no VILT")

                log("Tabela do VILT encontrada")
                return i + 1

    error("Não foi possível localizar a tabela de dados no VILT")

def parse_date(value):

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        value = value.strip()

        try:
            return datetime.strptime(value, "%Y/%m/%d")
        except ValueError:
            pass

        try:
            return datetime.strptime(value, "%d/%m/%Y")
        except ValueError:
            pass

    warn(f"Data inválida no VILT: {value}")
    return None

def parse_overwork(value):
    return str(value).strip().lower() in ['true', 'sim', '1', 'yes']

def filter_vilt():
    log("Processando VILT")

    wb = load_workbook(PATHS['vilt'])
    ws = find_vilt_sheet(wb)

    start_row = find_table_vilt(ws)

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        name = row[VILT_COLUMNS['Person']]
        if name not in INFOS:
            continue

        raw_date = parse_date(row[VILT_COLUMNS['Day']])

        if not raw_date:
            continue

        try:
            day = {
                'date': raw_date.strftime("%d/%m/%Y"),
                'hours': row[VILT_COLUMNS['Hours']] or 0,
                'overwork': parse_overwork(row[VILT_COLUMNS['Overwork']]),
                'activity': row[VILT_COLUMNS['Notes']] or ''
            }
        except Exception as e:
            warn(f"Erro ao processar linha do VILT ({name}): {e}")
            continue

        INFOS[name]['data'].setdefault(day['date'], []).append(day)

    log("VILT processado com sucesso")