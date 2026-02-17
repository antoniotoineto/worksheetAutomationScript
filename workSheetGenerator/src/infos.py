# =========================
# INFOS
# =========================

from openpyxl import load_workbook
from logger import error, log, warn
from config import PATHS

INFOS = {}

def fill_infos():
    log("Lendo infos.xlsx")

    wb = load_workbook(PATHS['infos'])
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        if not row[0]:
            continue

        name, reg, profile, email, squad, project, pm, cost_center, coordinator = row

        if not squad:
            warn(f"Recurso sem squad ignorado: {name}")
            continue

        INFOS[name.strip()] = {
            'registration': reg,
            'profile': profile,
            'email': email,
            'squad': str(squad).strip(),
            'project': project,
            'pm': pm,
            'cost_center': cost_center,
            'coordinator': coordinator,
            'data': {}
        }

    if not INFOS:
        error("Nenhum recurso válido encontrado no infos.xlsx")

    log(f"{len(INFOS)} recursos carregados")
