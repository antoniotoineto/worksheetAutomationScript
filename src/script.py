from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from copy import copy
import shutil
import os
import sys

# =========================
# CONFIGURAÇÕES
# =========================

def base_path():
    if getattr(sys, 'frozen', False):
        # Se estiver rodando como .exe
        return os.path.dirname(sys.executable)
    else:
        # Se estiver rodando como .py
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = base_path()

PATHS = {
    'template': os.path.join(BASE_DIR, 'base', 'template.xlsx'),
    'infos': os.path.join(BASE_DIR, 'base', 'infos.xlsx'),
    'vilt': os.path.join(BASE_DIR, 'vilt.xlsx'),
    'output': os.path.join(BASE_DIR, 'output')
}


SHEETS = {
    'base': 'sheet',
    'header': 'CABEÇALHO',
    'total': 'TOTALIZADOR'
}

VILT_SHEET_FALLBACK = 'Sheet1'

WEEKDAY = ['seg.', 'ter.', 'qua.', 'qui.', 'sex.', 'sab.', 'dom.']

ROW_IDENTIFIER = ['project', 'company']

VILT_COLUMNS = {
    'Person': None,
    'Day': None,
    'Hours': None,
    'Overwork': None,
    'Notes': None
}

INFOS = {}
DATE_RANGE = {'Start Date': '', 'End Date': ''}

os.makedirs(PATHS['output'], exist_ok=True)

# =========================
# LOG / ERRO
# =========================

def log(msg):
    print(f"[INFO] {msg}")

def warn(msg):
    print(f"[WARN] {msg}")

def error(msg):
    print(f"[ERROR] {msg}")
    sys.exit(1)

# =========================
# VALIDAÇÕES INICIAIS
# =========================

def validate_files():
    for k, path in PATHS.items():
        if k != 'output' and not os.path.exists(path):
            error(f"Arquivo não encontrado: {path}")
    log("Arquivos encontrados com sucesso")

# =========================
# INFOS
# =========================

def fill_infos():
    log("Lendo infos.xlsx")

    wb = load_workbook(PATHS['infos'])
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        if not row[0]:
            continue

        name, reg, profile, email, squad, project, pm, cost_center, coordinator = row

        if not squad:
            warn(f"Recurso sem squad ignorado: {name}")
            continue

        INFOS[name.strip()] = {
            'registration': reg,
            'profile': profile,
            'email': email,
            'squad': str(squad).strip(),
            'project': project,
            'pm': pm,
            'cost_center': cost_center,
            'coordinator': coordinator,
            'data': {}
        }

    if not INFOS:
        error("Nenhum recurso válido encontrado no infos.xlsx")

    log(f"{len(INFOS)} recursos carregados")

# =========================
# VILT
# =========================

def find_vilt_sheet(wb):
    if VILT_SHEET_FALLBACK in wb.sheetnames:
        return wb[VILT_SHEET_FALLBACK]
    return wb.active

def find_table_vilt(ws):
    log("Localizando tabela no VILT")

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue

        key = str(row[0]).strip().lower()

        if key.startswith('start'):
            start_date = parse_date(row[1])
            DATE_RANGE['Start Date'] = start_date

            year = start_date.year
            month = start_date.month

            if month == 12:
                next_month = 1
                year += 1
            else:
                next_month = month + 1

            DATE_RANGE['End Date'] = parse_date(date(year, next_month, 20))

        if len(row) > 1:
            if (
                str(row[0]).strip().lower() == ROW_IDENTIFIER[0]
                and str(row[1]).strip().lower() == ROW_IDENTIFIER[1]
            ):
                for idx, col in enumerate(row):
                    if col in VILT_COLUMNS:
                        VILT_COLUMNS[col] = idx

                if None in VILT_COLUMNS.values():
                    error("Colunas obrigatórias não encontradas no VILT")

                log("Tabela do VILT encontrada")
                return i + 1

    error("Não foi possível localizar a tabela de dados no VILT")

def parse_date(value):

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        value = value.strip()

        try:
            return datetime.strptime(value, "%Y/%m/%d")
        except ValueError:
            pass

        try:
            return datetime.strptime(value, "%d/%m/%Y")
        except ValueError:
            pass

    warn(f"Data inválida no VILT: {value}")
    return None


def parse_overwork(value):
    return str(value).strip().lower() in ['true', 'sim', '1', 'yes']

def filter_vilt():
    log("Processando VILT")

    wb = load_workbook(PATHS['vilt'])
    ws = find_vilt_sheet(wb)

    start_row = find_table_vilt(ws)

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        name = row[VILT_COLUMNS['Person']]
        if name not in INFOS:
            continue

        raw_date = parse_date(row[VILT_COLUMNS['Day']])

        if not raw_date:
            continue

        try:
            day = {
                'date': raw_date.strftime("%d/%m/%Y"),
                'weekday': WEEKDAY[raw_date.weekday()],
                'hours': row[VILT_COLUMNS['Hours']] or 0,
                'overwork': parse_overwork(row[VILT_COLUMNS['Overwork']]),
                'activity': row[VILT_COLUMNS['Notes']] or ''
            }
        except Exception as e:
            warn(f"Erro ao processar linha do VILT ({name}): {e}")
            continue

        INFOS[name]['data'].setdefault(day['date'], []).append(day)

    log("VILT processado com sucesso")

# =========================
# PLANILHAS PORTO
# =========================

def copy_style(ws, rows):
    base = ws[10]
    for r in range(10, rows):
        for c, ref in enumerate(base, start=1):
            cell = ws.cell(r, c)
            cell.font = copy(ref.font)
            cell.border = copy(ref.border)
            cell.fill = copy(ref.fill)
            cell.number_format = copy(ref.number_format)
            cell.alignment = copy(ref.alignment)

def fill_resource(ws, name, activities):
    info = INFOS[name]

    ws['B1'] = DATE_RANGE['Start Date']
    ws['C1'] = DATE_RANGE['End Date']
    ws['B4'] = name
    ws['A7'] = info['registration']
    ws['B7'] = info['email']
    ws['C7'] = info['profile']
    ws['D1'] = info['pm']

    row = 10
    last_date = None

    for a in activities:
        ws.cell(row, 1, a['date'])
        ws.cell(row, 2, a['weekday'])
        ws.cell(row, 3, a['hours'])
        ws.cell(row, 4, 'Sim' if a['overwork'] else 'Não')
        ws.cell(row, 5, a['hours'] * 2 if a['overwork'] else a['hours'])
        ws.cell(row, 6, a['activity'])
        last_date = parse_date(a['date'])
        row += 1

    if not last_date:
        last_date = DATE_RANGE['Start Date'] - timedelta(days=1)

    current_date = last_date + timedelta(days=1)
    end_date = DATE_RANGE['End Date']

    while current_date <= end_date:

        if current_date.weekday() < 5:  # 0=segunda, 6=domingo

            ws.cell(row, 1, current_date.strftime("%d/%m/%Y"))
            ws.cell(row, 2, WEEKDAY[current_date.weekday()])
            ws.cell(row, 3, 8)
            ws.cell(row, 4, 'Não')
            ws.cell(row, 5, 8)
            ws.cell(row, 6, 'Previsão de horas')

            row += 1

        current_date += timedelta(days=1)

    return info['profile'], row

def fill_header(ws, info):
    ws['B5'] = DATE_RANGE['Start Date']
    ws['C5'] = DATE_RANGE['End Date']
    ws['B7'] = info['pm']
    ws['B10'] = info['cost_center']
    ws['B12'] = info['coordinator']

def fill_total_infos(ws, info):
    ws['K3'] = info['cost_center']
    ws['O3'] = info['coordinator']  
    ws['A3'] = info['project']   

def fill_total_hours(ws, name, profile):

    sheet_reference = f"'{name}'!D7"

    for row in range(3, 8):

        profile_cell = ws.cell(row, 4)  # Coluna D
        hours_cell = ws.cell(row, 8)    # Coluna H

        # 🔹 Caso 1: perfil já existe → somar referência
        if profile_cell.value == profile:

            if hours_cell.value:
                hours_cell.value = f"{hours_cell.value}+{sheet_reference}"
            else:
                hours_cell.value = f"={sheet_reference}"

            return

        # 🔹 Caso 2: primeira linha vazia ("-")
        elif profile_cell.value == "-":

            profile_cell.value = profile
            hours_cell.value = f"={sheet_reference}"

            return

def group_by_squad():
    squads = {}
    for name, info in INFOS.items():
        squads.setdefault(info['squad'], []).append(name)
    return squads

def create_files():
    squads = group_by_squad()

    log(f"{len(squads)} squads identificadas")

    for squad, members in squads.items():
        log(f"Gerando planilha da squad: {squad}")

        output = f"{PATHS['output']}/{squad}.xlsx"
        shutil.copyfile(PATHS['template'], output)

        wb = load_workbook(output)

        if SHEETS['base'] not in wb.sheetnames:
            error("Aba base 'sheet' não encontrada no template")

        base_ws = wb[SHEETS['base']]

        fill_header(wb[SHEETS['header']], INFOS[members[0]])
        fill_total_infos(wb[SHEETS['total']], INFOS[members[0]])
        
        for name in members:
            ws = wb.copy_worksheet(base_ws)
            ws.title = name

            activities = [
                d for days in INFOS[name]['data'].values() for d in days
            ]

            profile, total_rows = fill_resource(ws, name, activities)
            copy_style(ws, total_rows)
            fill_total_hours(wb[SHEETS['total']], name, profile)


        wb.remove(base_ws)
        wb.save(output)

    log("Processo finalizado com sucesso")

# =========================
# MAIN
# =========================

if __name__ == '__main__':
    validate_files()
    fill_infos()
    filter_vilt()
    create_files()
